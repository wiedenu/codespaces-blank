#!/usr/bin/env python3
"""
MKB-6238 — Henry Schein $0 Invoice Email sender.

Rebuild (2026-07-07) of the original script lost in the May 20 workspace delete.
Sends via HubSpot's Single-Send Transactional Email API using a saved email ID.

Behavior mirrors the April 2026 run:
  - Reads a recipient list from an .xlsx export.
  - Main pass sends to every clean address, ONE at a time, with logging.
  - Rows with bad email data (multiple addresses, trailing semicolons, malformed)
    are NOT guessed at -- they're flagged and written to a resend workbook so you
    can correct them and run a second pass, exactly like last time (22 of 1,304).

Usage:
  # 1. Validate only -- no sends. Reports bad rows and writes the resend workbook.
  python3 mkb6238-rebate-sender.py "OPS - Henry Schein - 070726.xlsx" --dry-run

  # 2. Live send of the clean rows.
  python3 mkb6238-rebate-sender.py "OPS - Henry Schein - 070726.xlsx" --send

  # 3. After correcting addresses, run the resend workbook the same way.
  python3 mkb6238-rebate-sender.py "OPS - Henry Schein Resend - 070726.xlsx" --send

Auth:
  Set a HubSpot private-app token (needs the `transactional-email` scope) in the
  environment before running:
      export HUBSPOT_PRIVATE_APP_TOKEN="pat-na1-..."
  (HUBSPOT_ACCESS_TOKEN / HUBSPOT_TOKEN are also accepted.)
"""

import argparse
import json
import os
import re
import sys
import time
from collections import Counter
from datetime import datetime

import openpyxl
import requests

# --- Config -----------------------------------------------------------------

EMAIL_ID = 216664065950  # HubSpot saved single-send email for the $0 invoice
SEND_ENDPOINT = "https://api.hubapi.com/marketing/v3/transactional/single-email/send"

# Column whose header contains this (case-insensitive) is treated as the address.
EMAIL_HEADER_HINT = "email"

# The saved email (216664065950) personalizes with these contact.* tokens, so they
# must be sent as `contactProperties`. NOTE: contactProperties are written onto the
# contact record in HubSpot, not just into the email.
#   key   = HubSpot property internal name (the part after "contact." in the token)
#   value = column header in the export that holds that value
# If the export headers differ, update the right-hand values to match.
CONTACT_PROPERTY_MAP = {
    "smg_henry_schein_description": "Description",
    "pm_contract_id": "ApplicationID",     # 'Application Clean' is the same value as a number
    "pm_lessee_name": "Billing Name",
}

# Politeness / rate limiting.
SLEEP_BETWEEN_SENDS = 0.2   # seconds
MAX_RETRIES = 4             # on 429 / 5xx
BACKOFF_BASE = 2.0          # seconds, exponential

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


# --- Helpers ----------------------------------------------------------------

def get_token():
    for var in ("HUBSPOT_PRIVATE_APP_TOKEN", "HUBSPOT_ACCESS_TOKEN", "HUBSPOT_TOKEN"):
        tok = os.environ.get(var)
        if tok:
            return tok
    sys.exit(
        "ERROR: No HubSpot token found. Set one of "
        "HUBSPOT_PRIVATE_APP_TOKEN / HUBSPOT_ACCESS_TOKEN / HUBSPOT_TOKEN "
        "(private app with the 'transactional-email' scope)."
    )


def load_rows(path):
    """
    Return (headers, list-of-row-dicts) from the first worksheet.
    Auto-detects the header row: exports sometimes carry preamble/title rows above
    the real column headers (the June '26 file has two). The header row is the first
    one within the first 15 rows containing a cell that is exactly 'email'.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, values in enumerate(all_rows[:15]):
        if any(str(v).strip().lower() == EMAIL_HEADER_HINT for v in values if v is not None):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 0  # fall back to first row

    headers = [str(h).strip() if h is not None else "" for h in all_rows[header_idx]]
    rows = []
    for values in all_rows[header_idx + 1:]:
        if values is None or all(v is None for v in values):
            continue
        row = {}
        for h, v in zip(headers, values):
            if h:
                row[h] = "" if v is None else str(v).strip()
        rows.append(row)
    return headers, rows


def find_email_column(headers):
    for h in headers:
        if EMAIL_HEADER_HINT in h.lower():
            return h
    sys.exit(
        f"ERROR: No column containing '{EMAIL_HEADER_HINT}' found. "
        f"Headers were: {headers}"
    )


def classify_emails(raw):
    """
    Return (list_of_clean_addresses, reason).

    A cell may hold multiple addresses separated by ; or , -- we split and send to
    EACH (per John, 7/7/26). Every split piece must be a valid address; if any piece
    is invalid the whole cell is flagged 'malformed' for human correction (this also
    catches typos like 'drgme45@msn,.com', which are one bad address, not two).
    """
    if not raw or not raw.strip():
        return [], "empty"
    val = raw.strip().strip(";, ").strip()
    parts = [p.strip() for p in re.split(r"[;,]", val) if p.strip()]
    if not parts:
        return [], "empty"
    for p in parts:
        if not EMAIL_RE.match(p):
            return [], "malformed"
    # De-dupe identical addresses within the same cell, preserving order.
    seen, out = set(), []
    for a in parts:
        if a.lower() not in seen:
            seen.add(a.lower())
            out.append(a)
    return out, "ok"


def send_one(session, token, email_id, email, contact_props):
    payload = {
        "emailId": email_id,
        "message": {"to": email},
    }
    if contact_props:
        payload["contactProperties"] = contact_props
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    for attempt in range(1, MAX_RETRIES + 1):
        resp = session.post(SEND_ENDPOINT, headers=headers, data=json.dumps(payload), timeout=30)
        if resp.status_code in (200, 201):
            return True, resp.status_code, resp.text
        if resp.status_code == 429 or resp.status_code >= 500:
            wait = BACKOFF_BASE * (2 ** (attempt - 1))
            # Honor Retry-After if present.
            ra = resp.headers.get("Retry-After")
            if ra:
                try:
                    wait = float(ra)
                except ValueError:
                    pass
            time.sleep(wait)
            continue
        return False, resp.status_code, resp.text
    return False, resp.status_code, resp.text


def write_resend_workbook(path, headers, bad_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Needs Correction"
    ws.append(headers + ["_reason"])
    for row, reason in bad_rows:
        ws.append([row.get(h, "") for h in headers] + [reason])
    wb.save(path)


# --- Main -------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="MKB-6238 Henry Schein $0 invoice sender")
    ap.add_argument("infile", help="Recipient .xlsx export")
    mode = ap.add_mutually_exclusive_group(required=True)
    mode.add_argument("--dry-run", action="store_true",
                      help="Validate only; write resend workbook; send nothing")
    mode.add_argument("--send", action="store_true", help="Live send to clean rows")
    ap.add_argument("--email-id", type=int, default=EMAIL_ID,
                    help=f"Override HubSpot email ID (default {EMAIL_ID})")
    args = ap.parse_args()
    email_id = args.email_id

    stamp = datetime.now().strftime("%m%d%y")
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    base = os.path.splitext(os.path.basename(args.infile))[0]

    headers, rows = load_rows(args.infile)
    email_col = find_email_column(headers)

    # Verify every mapped personalization column is present in the export.
    missing_cols = [h for h in CONTACT_PROPERTY_MAP.values() if h not in headers]
    if missing_cols:
        print("WARNING: these mapped columns are missing from the export and will "
              f"send blank tokens: {missing_cols}")
        print(f"         export headers: {headers}")

    clean, bad = [], []
    blank_token_rows = 0
    split_rows = 0
    for row in rows:
        addrs, reason = classify_emails(row.get(email_col, ""))
        if not addrs:
            bad.append((row, reason))
            continue
        contact_props = {}
        for prop_name, header in CONTACT_PROPERTY_MAP.items():
            val = row.get(header, "")
            if val:
                contact_props[prop_name] = val
        if len(contact_props) < len(CONTACT_PROPERTY_MAP):
            blank_token_rows += 1
        if len(addrs) > 1:
            split_rows += 1
        for addr in addrs:
            clean.append((addr, contact_props))

    print(f"Loaded {len(rows)} rows from {args.infile}")
    print(f"  email column: '{email_col}'")
    print(f"  personalization -> contactProperties: {list(CONTACT_PROPERTY_MAP)}")
    print(f"  clean sends: {len(clean)}    flagged rows: {len(bad)}")
    if split_rows:
        print(f"  ({split_rows} rows had multiple addresses -> split into separate sends)")
    if blank_token_rows:
        print(f"  WARNING: {blank_token_rows} clean rows have >=1 blank personalization "
              "value (token will render empty for those recipients)")

    # Duplicate checks. Same email + same agreement = a true double-send (bad).
    # Same email across different agreements = likely legit (one contact, multiple
    # agreements) -- reported as informational, not removed.
    email_counts = Counter(addr for addr, _ in clean)
    dup_emails = {e: c for e, c in email_counts.items() if c > 1}
    pair_counts = Counter((addr, props.get("pm_contract_id", "")) for addr, props in clean)
    exact_dupes = {k: c for k, c in pair_counts.items() if c > 1}
    if dup_emails:
        print(f"  INFO: {len(dup_emails)} email(s) appear on multiple rows "
              "(may be legit -- one contact with multiple agreements):")
        for e, c in list(dup_emails.items())[:5]:
            print(f"        {e}  x{c}")
        if len(dup_emails) > 5:
            print(f"        ...and {len(dup_emails) - 5} more")
    if exact_dupes:
        print(f"  WARNING: {len(exact_dupes)} EXACT duplicate row(s) "
              "(same email AND same agreement) -- these would send the same invoice twice:")
        for (e, cid), c in list(exact_dupes.items())[:10]:
            print(f"        {e}  agreement={cid}  x{c}")

    if bad:
        by_reason = {}
        for _, r in bad:
            by_reason[r] = by_reason.get(r, 0) + 1
        print("  flagged breakdown:", dict(by_reason))

    # Always write the resend workbook when there are flagged rows.
    if bad:
        resend_path = f"OPS - Henry Schein Resend - {stamp}.xlsx"
        write_resend_workbook(resend_path, headers, bad)
        print(f"  -> wrote {len(bad)} flagged rows to: {resend_path}")

    if args.dry_run:
        print("\nDRY RUN -- nothing sent. Review the flagged rows, correct addresses,")
        print("then run --send on the corrected workbook.")
        return

    # Live send.
    token = get_token()
    log_path = f"mkb6238-send-{ts}.log"
    ok = fail = 0
    session = requests.Session()
    with open(log_path, "w") as log:
        log.write(f"MKB-6238 send @ {ts}  emailId={email_id}  file={args.infile}\n")
        log.write(f"clean={len(clean)}  flagged-not-sent={len(bad)}\n")
        for i, (addr, props) in enumerate(clean, 1):
            success, code, body = send_one(session, token, email_id, addr, props)
            status = "OK " if success else "ERR"
            log.write(f"{status} {code} {addr}\n")
            log.flush()
            if not success:
                log.write(f"    body: {body[:500]}\n")
                fail += 1
            else:
                ok += 1
            if i % 50 == 0 or i == len(clean):
                print(f"  sent {i}/{len(clean)}  (ok={ok} err={fail})")
            time.sleep(SLEEP_BETWEEN_SENDS)
        log.write(f"\nSUMMARY: ok={ok}  err={fail}  flagged-not-sent={len(bad)}  "
                  f"total-rows={len(clean) + len(bad)}\n")

    print(f"\nDone. ok={ok}  err={fail}  flagged-not-sent={len(bad)}")
    print(f"Log: {log_path}")
    if fail:
        print("Some sends errored -- check the log; those addresses may need a resend.")


if __name__ == "__main__":
    main()
